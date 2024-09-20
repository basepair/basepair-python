#!/usr/bin/env python

from openpyxl import load_workbook


# read info from an excel file and create samples
def create_samples(bp, filename):
    wb = load_workbook(filename)
    ws = wb.get_sheet_by_name('Sheet1')

    header = [cell.value for cell in ws.rows[0]]
    for k in xrange(1, ws.max_row):
        vals = [cell.value for cell in ws.rows[k]]
        rec = dict(zip(header, vals))

        data = {
            'name': rec['Name'],
            'genome': rec['Genome'],
            'datatype': rec['Data type'],
        }

        # # existing record, skip
        if rec['Sample Id']:
            # continue
            data['id'] = rec['Sample Id']

        if 'Workflow' in rec:
            data['default_workflow'] = rec['Workflow']
        if 'Platform' in rec:
            data['platform'] = rec['Platform']
        if 'Filepath 1' in rec:
            data['filepath1'] = rec['Filepath 1']
        if 'Filepath 2' in rec:
            data['filepath2'] = rec['Filepath 2']
        if 'S3path' in rec:
            data['s3path'] = rec['S3path']
        if 'S3path_pe' in rec:
            data['s3path_pe'] = rec['S3path_pe']

        sample_id = bp.create_sample(data, upload=False)
        ws.rows[k][0].value = sample_id

    wb.save(filename)


# read info from an excel file and create samples
def start_analyses(bp, filename):
    wb = load_workbook(filename)
    ws = wb.get_sheet_by_name('Sheet1')

    header = [cell.value for cell in ws.rows[0]]
    for k in xrange(1, ws.max_row):
        vals = [cell.value for cell in ws.rows[k]]
        rec = dict(zip(header, vals))

        if rec['Analysis Id']:
            continue

        sample_id = int(rec['Sample Id'])
        workflow_id = int(rec['Workflow'])

        analysis_id = bp.create_analysis(
            workflow_id=workflow_id, sample_id=sample_id)
        ws.rows[k][1].value = analysis_id

        # break

    wb.save(filename)


def download_raw(bp, filename):
    wb = load_workbook(filename)
    ws = wb.worksheets[0]

    header = [cell.value for cell in ws.rows[0]]
    for k in xrange(1, ws.max_row):
        vals = [cell.value for cell in ws.rows[k]]
        rec = dict(zip(header, vals))

        sample = bp.get_info('samples', rec['Id'])
        keys = [sample['s3path'], sample['s3path_pe']]
        for key in keys:
            if not key:
                continue

            filename = '{}-{}/{}'.format(
                rec['Id'], rec['Name'].replace(' ', '_'),
                key.rsplit('/', 1)[1])

            bp.get_file(key, filename)
